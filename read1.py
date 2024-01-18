from flask import Flask, render_template, request, flash, Response, session, send_from_directory
from flask_wtf import FlaskForm
from wtforms import SubmitField
from werkzeug.utils import secure_filename
import csv
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error
from sklearn.impute import SimpleImputer
import seaborn as sns



app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads/'
ALLOWED_EXTENSIONS = {'csv'}

class UploadForm(FlaskForm):
    submit = SubmitField('Upload')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def predict_missing_values(df, parameter):
    # Exclude 'Time' column from features
    X = df.drop([parameter, 'Time'], axis=1)
    y = df[parameter]

    # Check if there are missing values or zeros in the target variable
    if (y.isnull() | (y == 0)).any():
        # Consider 0 as a missing value
        y.replace(0, float('nan'), inplace=True)

        # Split the data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # Impute missing values in the target variable
        imputer = SimpleImputer(strategy='mean')
        y_train_imputed = imputer.fit_transform(y_train.values.reshape(-1, 1)).ravel()
        y_test_imputed = imputer.transform(y_test.values.reshape(-1, 1)).ravel()

        # Train a machine learning model (Random Forest Regressor in this example)
        model = RandomForestRegressor()
        model.fit(X_train, y_train_imputed)

        # Use the trained model to predict missing values in the specified parameter
        X_missing = df[(df[parameter].isnull()) | (df[parameter] == 0)].drop([parameter, 'Time'], axis=1)

        # Check if there are any rows in X_missing
        if not X_missing.empty:
            predicted_values = model.predict(X_missing)

            # Fill in the missing values in the original dataframe
            df.loc[(df[parameter].isnull()) | (df[parameter] == 0), parameter] = predicted_values
        else:
            print(f"No rows with missing values or zeros in {parameter} to predict.")
    else:
        print(f"No missing values or zeros in {parameter}.")

    return df

def generate_metrics_plot(df, selected_parameter):
    # Your existing code to calculate metrics and plot
    n = df[selected_parameter]
    bias_temp_500m = n - n.mean()
    std_temp_500m = n.std()
    rms_temp_500m = np.sqrt(((bias_temp_500m) ** 2).mean())
    df['Time'] = pd.to_datetime(df['Time'], format='%d.%m.%Y %H:%M:%S')
    monthly_metrics = pd.DataFrame(index=pd.to_datetime(df['Time']).dt.to_period('M'))
    monthly_metrics['Bias'] = n.groupby(monthly_metrics.index).mean()
    monthly_metrics['Std Dev'] = n.groupby(monthly_metrics.index).std()
    monthly_metrics['RMS'] = n.groupby(monthly_metrics.index).apply(lambda x: np.sqrt(((x - x.mean()) ** 2).mean()))
    monthly_metrics['Nb Obs'] = n.groupby(monthly_metrics.index).count()

    metrics = ['Bias', 'Std Dev', 'RMS']
    fig, axes = plt.subplots(nrows=3, figsize=(12, 18), sharex=True)

    for i, metric in enumerate(metrics):
        ax = axes[i]
        ax.plot(monthly_metrics.index.astype(str), monthly_metrics[metric], label=metric, color='darkblue', linewidth=3)

        # Add upper and lower range lines in red
        upper_range = n.mean() + 2 * std_temp_500m
        lower_range = n.mean() - 2 * std_temp_500m
        ax.axhline(y=upper_range, color='red', label='Upper Range')
        ax.axhline(y=lower_range, color='red', label='Lower Range')

        # Label months under all three graphs
        ax.set_xticks(monthly_metrics.index.astype(str))
        ax.set_xticklabels(monthly_metrics.index.astype(str), rotation=45, ha='right')

        # Add dots at turning points
        sns.scatterplot(data=monthly_metrics, x=monthly_metrics.index.astype(str), y=metric, ax=ax, color='darkblue', s=100, zorder=5)

        ax.set_title(f'{selected_parameter} {metric} by Month')
        ax.set_xlabel('Month')
        ax.set_ylabel(metric)
        ax.legend()
        ax.grid(True, alpha=0.5,linewidth = 0.7,color = 'grey')

        ax_nb_obs = ax.twinx()
        ax_nb_obs.plot(monthly_metrics.index.astype(str), monthly_metrics['Nb Obs'], linestyle='--', color='purple', label='Nb Obs', linewidth=2)
        ax_nb_obs.set_ylabel('Nb Obs')
        ax_nb_obs.legend(loc='upper right')

    plt.tight_layout()

    # Save the plot to a file and return the file path
    plot_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f'metrics_plot_{selected_parameter}.png')
    plt.savefig(plot_file_path)
    plt.close()  # Close the plot to free up resources
    return plot_file_path

def flag_data(rows, parameter):
    headers = rows[0]
    flagged_data = []
    

    for i, row in enumerate(rows[1:], 1):  # Starting at 1 for row numbers
        param_index = headers.index(parameter)
        value = float(row[param_index])
        longitude = float(row[headers.index("Longitude")])

        flag_msg = []


        # Check for Impossible location
        if not (94 <= longitude <= 94.5):
            flag_msg.append("Impossible location")

        if value == 0:  # Check for missing values (0)
            
            flag_msg.append("Missing Value")

        elif parameter == "Temperature_500m":
            # Check outside range
            if not (0 <= value <= 35):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 3:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")
        elif parameter == "WindGust":
            # Check outside range
            if not (0 <= value <= 70):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 7:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")
        elif parameter == "WindSpeed":
            # Check outside range
            if not (0 <= value <= 70):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 7:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "Air_Temp":
            # Check outside range
            if not (20<= value <= 35):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 5:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")
        
        elif parameter == "Conductivity_0001m":
            # Check outside range
            if not (20 <= value <= 70):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 2:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "Precipitation":
            # Check outside range
            if not (0 <= value <= 50):
                flag_msg.append("Outside Range")
            
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "pressure_500m":
            # Check outside range
            if not (45 <= value <= 55):
                flag_msg.append("Outside Range")
            
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "Rel_Hum":
            # Check outside range
            if not (50 <= value <= 100):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 8:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "WindDirection":
            # Check outside range
            if not (0 <= value <= 360):
                flag_msg.append("Outside Range")
            
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        elif parameter == "SLP":
            # Check outside range
            if not (990 <= value <= 1020):
                flag_msg.append("Outside Range")
            # Check spikes
            if i > 1:
                prev_value = float(rows[i-1][param_index])
                if abs(value - prev_value) > 5:  # Assuming 40% spike check
                    flag_msg.append("Spike Detected")
            # Check stuck values
            if i > 3 and all(value == float(rows[j][param_index]) for j in range(i-2, i)):
                flag_msg.append("Stuck value detected")

        if flag_msg:
            flagged_data.append([i, row[param_index], ", ".join(flag_msg)])  # row number, flagged value, flagged issue
        summary = {
        'Total': len(rows) - 1,
        'Flagged': len(flagged_data),
        'Impossible Location': sum(1 for _, _, issues in flagged_data if "Impossible location" in issues),
        'Outside Range': sum(1 for _, _, issues in flagged_data if "Outside Range" in issues),
        'Spike Detected': sum(1 for _, _, issues in flagged_data if "Spike Detected" in issues),
        'Stuck Value': sum(1 for _, _, issues in flagged_data if "Stuck value detected" in issues),
        'No of Missing Values': sum(1 for _, _, issues in flagged_data if "Missing Value" in issues)  # Include missing values count in the summary
    }

    return flagged_data, summary

def get_col_letter(col_num):
    string = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        string = chr(65 + remainder) + string
    return string

def save_report_to_excel(flagged_data, rows, parameter):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    color_map = {
        "Impossible location": ("FFFF0000", "❗"),
        "Outside Range": ("FFFFFF00", "⚠"),
        "Spike Detected": ("FF00FF00", "⬆"),
        "Stuck value detected": ("FF0000FF", "🔄"),
    }

    start_row = len(color_map) + 4
    ws1.cell(row=start_row, column=1, value="Summary")
    start_row += 1
    for key, value in session.get('summary').items():
        ws1.cell(row=start_row, column=1, value=key)
        ws1.cell(row=start_row, column=2, value=value)
        start_row += 1

    ws1.cell(row=1, column=1, value="Color & Symbol Key")
    ws1.cell(row=1, column=2, value="Reason")
    for index, (reason, (color_code, symbol)) in enumerate(color_map.items(), start=2):
        cell_color = ws1.cell(row=index, column=1)
        cell_color.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        cell_color.value = symbol

        cell_reason = ws1.cell(row=index, column=2)
        cell_reason.value = reason

    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 25

    ws2 = wb.create_sheet(title="Flagged Data")
    for col_num, header in enumerate(rows[0], 1):
        col_letter = get_col_letter(col_num)
        cell = ws2['{}1'.format(col_letter)]
        cell.value = header

    flagged_dict = {row_index: (value, issues) for row_index, value, issues in flagged_data}
    for row_num, row in enumerate(rows[1:], 2):
        for col_num, value in enumerate(row, 1):
            col_letter = get_col_letter(col_num)
            cell = ws2['{}{}'.format(col_letter, row_num)]

            if row_num - 1 in flagged_dict and rows[0][col_num - 1] == parameter:
                _, issues = flagged_dict[row_num - 1]
                symbols = []

                # Check other issues and update cell color accordingly
                for issue in issues.split(", "):
                    if issue in color_map :
                        cell.fill = PatternFill(start_color=color_map[issue][0], end_color=color_map[issue][0], fill_type="solid")
                        symbols.append(color_map[issue][1])

                cell.value = value + " " + ' '.join(symbols)
            else:
                cell.value = value

    ws3 = wb.create_sheet(title="Graphs")

    # Add the graphs to the "Graphs" sheet
    img_path = os.path.join(app.config['UPLOAD_FOLDER'], f'metrics_plot_{parameter}.png')
    img = Image(img_path)

    ws3.add_image(img, 'A1')            

    report_path = os.path.join(app.config['UPLOAD_FOLDER'], f'report_{parameter}.xlsx')
    wb.save(report_path)

    return report_path



@app.route('/', methods=['GET', 'POST'])
def upload_file():
    form = UploadForm()
    summary = None
    flags = []
    headers = []
    selected_parameter = request.form.get("parameter", "Temperature_500m")  # Default to "Temperature_500m"
    filepath = ""
    plot_image_metrics = None  # Initialize to None

    if 'file' in request.files and allowed_file(request.files['file'].filename):
        file = request.files['file']
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
        file.save(filepath)
    elif 'filepath' in request.form:
        filepath = request.form['filepath']

    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            rows = list(reader)
            headers = rows[0]

    flagged_rows = []
    if form.validate_on_submit():
        flagged_rows, summary_data = flag_data(rows, selected_parameter)
        session['summary'] = summary_data
        session['parameter'] = selected_parameter
        session.modified = True

        flags.extend(flagged_rows)
        summary = {
            "Total": len(rows) - 1,
            "Flagged": len(flagged_rows)
        }

        df = predict_missing_values(pd.read_csv(filepath), selected_parameter)
        updated_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'updated_data_{selected_parameter}.csv')
        df.to_csv(updated_filepath, index=False)
        session['updated_filepath'] = updated_filepath

        # Include the code to generate the metrics plot
        plot_image_metrics = generate_metrics_plot(df, selected_parameter)

    return render_template('upload.html', form=form, summary=summary, flags=flags, headers=headers, parameter=selected_parameter, filepath=filepath, plot_image_metrics=plot_image_metrics)



@app.route('/download-report')
def download_updated_data():
    updated_filepath = session.get('updated_filepath')
    selected_parameter = session.get('parameter', 'Temperature_500m')  # Use the selected parameter from the session

    if not updated_filepath or not os.path.exists(updated_filepath):
        return "No updated data available", 404

    # Read the updated CSV file using csv.reader and convert it to a list
    with open(updated_filepath, 'r') as updated_file:
        updated_reader = csv.reader(updated_file)
        updated_rows = list(updated_reader)

    # Get the flagged data and summary for the downloaded file using the selected parameter
    flagged_rows, _ = flag_data(updated_rows, selected_parameter)

    # Save the report to Excel
    report_path = save_report_to_excel(flagged_rows, updated_rows, selected_parameter)

    return send_from_directory(directory=app.config['UPLOAD_FOLDER'], path=os.path.basename(report_path), as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(debug=True)


